# ///////////////////////////////////////////////////////////////
# _formatters - Closed-file Excel formatting via openpyxl
# Project: EzXl
# ///////////////////////////////////////////////////////////////

"""
ExcelFormatter — fluent API for formatting Excel files without COM.

Operates exclusively on **closed** workbook files using openpyxl. All
formatting operations are buffered internally and applied in a single
pass when ``save()`` is called.

This module has no dependency on pywin32 or a running Excel process.
"""

from __future__ import annotations

# ///////////////////////////////////////////////////////////////
# IMPORTS
# ///////////////////////////////////////////////////////////////
# Standard library imports
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

# Third-party imports
import openpyxl
from ezplog.lib_mode import get_logger, get_printer
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

# ///////////////////////////////////////////////////////////////
# CONSTANTS
# ///////////////////////////////////////////////////////////////

logger = get_logger(__name__)
printer = get_printer()

# ///////////////////////////////////////////////////////////////
# INTERNAL DATA STRUCTURES
# ///////////////////////////////////////////////////////////////
# Each operation is stored as a dataclass so that save() can apply them
# in insertion order without repeatedly touching the openpyxl workbook.


@dataclass
class _ColumnWidthOp:
    col: str
    width: float


@dataclass
class _RowHeightOp:
    row: int
    height: float


@dataclass
class _FontOp:
    ref: str
    bold: bool
    italic: bool
    size: int | None
    color: str | None  # hex without '#'


@dataclass
class _FillOp:
    ref: str
    color: str  # hex without '#'


@dataclass
class _BorderOp:
    ref: str
    style: str


@dataclass
class _AlignOp:
    ref: str
    horizontal: str | None
    vertical: str | None
    wrap: bool


_Operation = _ColumnWidthOp | _RowHeightOp | _FontOp | _FillOp | _BorderOp | _AlignOp

# ///////////////////////////////////////////////////////////////
# CLASSES
# ///////////////////////////////////////////////////////////////


class ExcelFormatter:
    """Fluent formatter for closed Excel workbook files.

    All formatting operations are buffered and applied in a single write
    pass when ``save()`` is called. The workbook is opened with openpyxl
    only at save time, minimising I/O overhead.

    The API is intentionally flat: no sheet selector is exposed here. The
    formatter operates on the **active sheet** of the workbook. Consumer
    libraries that need multi-sheet formatting should instantiate one
    ``ExcelFormatter`` per sheet operation.

    Args:
        path: Path to an existing ``.xlsx`` workbook file.

    Raises:
        FileNotFoundError: If ``path`` does not exist.
        ImportError: If openpyxl is not installed.

    Example:
        >>> (
        ...     ExcelFormatter("report.xlsx")
        ...     .column_width("A", 20)
        ...     .font("A1", bold=True, size=14, color="FFFFFF")
        ...     .fill("A1", "4F81BD")
        ...     .save()
        ... )
    """

    # ///////////////////////////////////////////////////////////////
    # INIT
    # ///////////////////////////////////////////////////////////////

    def __init__(self, path: str | Path) -> None:
        self._path = Path(path).resolve()
        if not self._path.exists():
            raise FileNotFoundError(f"ExcelFormatter: file not found: {self._path}")
        # Buffer of pending operations applied in order at save().
        self._ops: list[_Operation] = []

    # ///////////////////////////////////////////////////////////////
    # PUBLIC METHODS — formatting operations (all return self for chaining)
    # ///////////////////////////////////////////////////////////////

    def column_width(self, col: str, width: float) -> ExcelFormatter:
        """Set the width of a column.

        Args:
            col: Column letter (e.g. ``"A"``, ``"BC"``).
            width: Column width in Excel character units.

        Returns:
            ExcelFormatter: ``self`` for method chaining.

        Example:
            >>> formatter.column_width("A", 20).column_width("B", 15)
        """
        self._ops.append(_ColumnWidthOp(col=col, width=width))
        return self

    def row_height(self, row: int, height: float) -> ExcelFormatter:
        """Set the height of a row.

        Args:
            row: 1-based row index.
            height: Row height in points.

        Returns:
            ExcelFormatter: ``self`` for method chaining.

        Example:
            >>> formatter.row_height(1, 30)
        """
        self._ops.append(_RowHeightOp(row=row, height=height))
        return self

    def font(
        self,
        ref: str,
        *,
        bold: bool = False,
        italic: bool = False,
        size: int | None = None,
        color: str | None = None,
    ) -> ExcelFormatter:
        """Apply font formatting to a cell or range.

        Args:
            ref: Cell or range address in A1 notation (e.g. ``"A1"`` or
                ``"A1:D1"``).
            bold: Apply bold weight. Defaults to ``False``.
            italic: Apply italic style. Defaults to ``False``.
            size: Font size in points. ``None`` leaves the size unchanged.
            color: Font colour as a 6-character hex string without ``#``
                (e.g. ``"FF0000"`` for red). ``None`` leaves colour unchanged.

        Returns:
            ExcelFormatter: ``self`` for method chaining.

        Example:
            >>> formatter.font("A1", bold=True, size=12, color="FF0000")
        """
        self._ops.append(
            _FontOp(ref=ref, bold=bold, italic=italic, size=size, color=color)
        )
        return self

    def fill(self, ref: str, color: str) -> ExcelFormatter:
        """Apply a solid background fill to a cell or range.

        Args:
            ref: Cell or range address in A1 notation.
            color: Background colour as a 6-character hex string without
                ``#`` (e.g. ``"4F81BD"`` for a medium blue).

        Returns:
            ExcelFormatter: ``self`` for method chaining.

        Example:
            >>> formatter.fill("A1:D1", "4F81BD")
        """
        self._ops.append(_FillOp(ref=ref, color=color))
        return self

    def border(self, ref: str, style: str = "thin") -> ExcelFormatter:
        """Apply a border to all edges of a cell or range.

        Args:
            ref: Cell or range address in A1 notation.
            style: Border style name as understood by openpyxl
                (e.g. ``"thin"``, ``"medium"``, ``"thick"``, ``"dashed"``).
                Defaults to ``"thin"``.

        Returns:
            ExcelFormatter: ``self`` for method chaining.

        Example:
            >>> formatter.border("A1:D5", style="thin")
        """
        self._ops.append(_BorderOp(ref=ref, style=style))
        return self

    def align(
        self,
        ref: str,
        *,
        horizontal: str | None = None,
        vertical: str | None = None,
        wrap: bool = False,
    ) -> ExcelFormatter:
        """Apply alignment to a cell or range.

        Args:
            ref: Cell or range address in A1 notation.
            horizontal: Horizontal alignment. Accepted values:
                ``"left"``, ``"center"``, ``"right"``, ``"fill"``,
                ``"justify"``, ``"centerContinuous"``, ``"distributed"``.
                ``None`` leaves the setting unchanged.
            vertical: Vertical alignment. Accepted values:
                ``"top"``, ``"center"``, ``"bottom"``, ``"justify"``,
                ``"distributed"``. ``None`` leaves the setting unchanged.
            wrap: Enable text wrapping. Defaults to ``False``.

        Returns:
            ExcelFormatter: ``self`` for method chaining.

        Example:
            >>> formatter.align("A1", horizontal="center", vertical="top", wrap=True)
        """
        self._ops.append(
            _AlignOp(ref=ref, horizontal=horizontal, vertical=vertical, wrap=wrap)
        )
        return self

    def save(self, dest: str | Path | None = None) -> None:
        """Apply all buffered operations and write the workbook.

        Args:
            dest: Destination path. Pass ``None`` to overwrite the source
                file in place. Parent directories of ``dest`` must exist.

        Raises:
            FormatterError: If any openpyxl operation fails.
            ImportError: If openpyxl is not installed.

        Example:
            >>> formatter.save()                        # overwrite source
            >>> formatter.save("output/report.xlsx")   # write to new path
        """
        from ..exceptions import FormatterError

        logger.debug(
            "ExcelFormatter.save: applying %d operations to '%s'.",
            len(self._ops),
            self._path,
        )

        try:
            wb = openpyxl.load_workbook(self._path)
        except Exception as exc:
            raise FormatterError(
                f"Failed to open workbook for formatting: {self._path} — {exc}",
                cause=exc,
            ) from exc

        try:
            ws = wb.active
            if ws is None:
                raise FormatterError(
                    f"Workbook '{self._path}' has no active sheet.",
                    cause=None,
                )

            for op in self._ops:
                if isinstance(op, _ColumnWidthOp):
                    ws.column_dimensions[op.col].width = op.width

                elif isinstance(op, _RowHeightOp):
                    ws.row_dimensions[op.row].height = op.height

                elif isinstance(op, _FontOp):
                    font_kwargs: dict[str, Any] = {
                        "bold": op.bold,
                        "italic": op.italic,
                    }
                    if op.size is not None:
                        font_kwargs["size"] = op.size
                    if op.color is not None:
                        font_kwargs["color"] = op.color
                    font_style = Font(**font_kwargs)
                    for cell in _iter_cells(ws, op.ref):
                        cell.font = font_style

                elif isinstance(op, _FillOp):
                    fill_style = PatternFill(fill_type="solid", fgColor=op.color)
                    for cell in _iter_cells(ws, op.ref):
                        cell.fill = fill_style

                elif isinstance(op, _BorderOp):
                    side = Side(border_style=cast(Any, op.style))
                    border_style = Border(left=side, right=side, top=side, bottom=side)
                    for cell in _iter_cells(ws, op.ref):
                        cell.border = border_style

                elif isinstance(op, _AlignOp):
                    align_kwargs: dict[str, Any] = {"wrap_text": op.wrap}
                    if op.horizontal is not None:
                        align_kwargs["horizontal"] = op.horizontal
                    if op.vertical is not None:
                        align_kwargs["vertical"] = op.vertical
                    align_style = Alignment(**align_kwargs)
                    for cell in _iter_cells(ws, op.ref):
                        cell.alignment = align_style

        except Exception as exc:
            raise FormatterError(
                f"Error applying formatting operation to '{self._path}': {exc}",
                cause=exc,
            ) from exc

        out_path = Path(dest).resolve() if dest is not None else self._path

        try:
            wb.save(str(out_path))
        except Exception as exc:
            raise FormatterError(
                f"Failed to save formatted workbook to '{out_path}': {exc}",
                cause=exc,
            ) from exc

        logger.debug("ExcelFormatter.save: written to '%s'.", out_path)
        printer.success(
            f"ExcelFormatter: formatting applied and saved to '{out_path}'."
        )


# ///////////////////////////////////////////////////////////////
# HELPERS
# ///////////////////////////////////////////////////////////////


def _iter_cells(ws: Any, ref: str):
    """Yield individual Cell objects from a worksheet reference.

    Handles both single-cell references (``"A1"``) and ranges (``"A1:D5"``).

    Args:
        ws: An openpyxl Worksheet object.
        ref: Cell or range address in A1 notation.

    Yields:
        openpyxl Cell objects.
    """
    cell_range = ws[ref]
    # openpyxl returns a single Cell for a single address, or a tuple-of-tuples
    # for a range reference.
    if hasattr(cell_range, "__iter__") and not hasattr(cell_range, "value"):
        # Range: iterate rows then cells.
        for row in cell_range:
            if hasattr(row, "__iter__") and not hasattr(row, "value"):
                yield from row
            else:
                yield row
    else:
        yield cell_range
