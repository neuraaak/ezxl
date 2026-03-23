# ///////////////////////////////////////////////////////////////
# test_formatters - EzXl closed-file formatter tests
# Project: EzXl
# ///////////////////////////////////////////////////////////////

"""
Unit tests for ``ezxl.io._formatters``.

These tests exercise both the successful formatting flow and the main
error branches raised by ``ExcelFormatter.save()``. All scenarios work on
closed files only and do not require a running Excel process.
"""

from __future__ import annotations

# ///////////////////////////////////////////////////////////////
# IMPORTS
# ///////////////////////////////////////////////////////////////
# Standard library imports
from pathlib import Path

# Third-party imports
import openpyxl
import pytest

# Local imports
from ezxl.exceptions import FormatterError
from ezxl.io import _formatters
from ezxl.io._formatters import ExcelFormatter, _iter_cells

# ///////////////////////////////////////////////////////////////
# TESTS — construction
# ///////////////////////////////////////////////////////////////


@pytest.mark.unit
def test_should_raise_filenotfound_when_formatter_source_is_missing(
    temp_dir: Path,
) -> None:
    """Verify that ``ExcelFormatter`` rejects a missing workbook path."""
    missing = temp_dir / "missing.xlsx"

    with pytest.raises(FileNotFoundError):
        ExcelFormatter(missing)


# ///////////////////////////////////////////////////////////////
# TESTS — helper iteration
# ///////////////////////////////////////////////////////////////


@pytest.mark.unit
def test_should_iter_cells_support_single_cell_and_range(sample_xlsx: Path) -> None:
    """Verify that ``_iter_cells`` yields cells for single refs and ranges."""
    workbook = openpyxl.load_workbook(sample_xlsx)
    worksheet = workbook.active

    assert worksheet is not None

    single = list(_iter_cells(worksheet, "A1"))
    range_cells = list(_iter_cells(worksheet, "A1:B2"))

    assert len(single) == 1
    assert single[0].coordinate == "A1"
    assert [cell.coordinate for cell in range_cells] == ["A1", "B1", "A2", "B2"]


# ///////////////////////////////////////////////////////////////
# TESTS — successful formatting
# ///////////////////////////////////////////////////////////////


@pytest.mark.unit
def test_should_apply_buffered_operations_when_saving_in_place(
    sample_xlsx: Path,
) -> None:
    """Verify that all buffered style and sizing operations are persisted."""
    (
        ExcelFormatter(sample_xlsx)
        .column_width("A", 22)
        .row_height(1, 30)
        .font("A1:C1", bold=True, italic=True, size=14, color="FF0000")
        .fill("A1:C1", "4F81BD")
        .border("A1:C1", style="thin")
        .align("A1:C1", horizontal="center", vertical="top", wrap=True)
        .save()
    )

    workbook = openpyxl.load_workbook(sample_xlsx)
    worksheet = workbook.active

    assert worksheet is not None
    assert worksheet.column_dimensions["A"].width == 22
    assert worksheet.row_dimensions[1].height == 30

    header_cell = worksheet["A1"]
    assert header_cell.font.bold is True
    assert header_cell.font.italic is True
    assert header_cell.font.size == 14
    assert str(header_cell.font.color.rgb).endswith("FF0000")
    assert header_cell.fill.fill_type == "solid"
    assert str(header_cell.fill.fgColor.rgb).endswith("4F81BD")
    assert header_cell.border.left.style == "thin"
    assert header_cell.border.right.style == "thin"
    assert header_cell.border.top.style == "thin"
    assert header_cell.border.bottom.style == "thin"
    assert header_cell.alignment.horizontal == "center"
    assert header_cell.alignment.vertical == "top"
    assert header_cell.alignment.wrap_text is True


@pytest.mark.unit
def test_should_save_formatted_workbook_to_new_destination(
    sample_xlsx: Path, temp_dir: Path
) -> None:
    """Verify that ``save(dest=...)`` writes to a new workbook path."""
    dest = temp_dir / "formatted.xlsx"

    ExcelFormatter(sample_xlsx).column_width("B", 18).save(dest)

    assert dest.exists()

    source_workbook = openpyxl.load_workbook(sample_xlsx)
    source_sheet = source_workbook.active
    assert source_sheet is not None

    dest_workbook = openpyxl.load_workbook(dest)
    dest_sheet = dest_workbook.active
    assert dest_sheet is not None

    assert source_sheet.column_dimensions["B"].width != 18
    assert dest_sheet.column_dimensions["B"].width == 18


# ///////////////////////////////////////////////////////////////
# TESTS — error handling
# ///////////////////////////////////////////////////////////////


@pytest.mark.unit
def test_should_raise_formatter_error_when_workbook_open_fails(
    sample_xlsx: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Verify that workbook-open failures are wrapped in ``FormatterError``."""

    def _raise_open_error(_path: Path) -> None:
        raise OSError("boom")

    monkeypatch.setattr(_formatters.openpyxl, "load_workbook", _raise_open_error)

    with pytest.raises(FormatterError, match="Failed to open workbook") as exc_info:
        ExcelFormatter(sample_xlsx).save()

    assert isinstance(exc_info.value.cause, OSError)


@pytest.mark.unit
def test_should_raise_formatter_error_when_workbook_has_no_active_sheet(
    sample_xlsx: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Verify that a workbook without an active sheet is rejected."""

    class _WorkbookWithoutActiveSheet:
        active = None

        def save(self, _dest: str) -> None:
            raise AssertionError("save() should not be called when active is None")

    monkeypatch.setattr(
        _formatters.openpyxl,
        "load_workbook",
        lambda _path: _WorkbookWithoutActiveSheet(),
    )

    with pytest.raises(FormatterError, match="has no active sheet"):
        ExcelFormatter(sample_xlsx).save()


@pytest.mark.unit
def test_should_raise_formatter_error_when_format_application_fails(
    sample_xlsx: Path,
) -> None:
    """Verify that invalid worksheet references are wrapped during formatting."""
    formatter = ExcelFormatter(sample_xlsx).font("not-a-ref", bold=True)

    with pytest.raises(FormatterError, match="Error applying formatting operation"):
        formatter.save()


@pytest.mark.unit
def test_should_raise_formatter_error_when_save_fails(
    sample_xlsx: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Verify that write failures are wrapped in ``FormatterError``."""
    workbook = openpyxl.load_workbook(sample_xlsx)

    def _raise_save_error(_dest: str) -> None:
        raise PermissionError("locked")

    monkeypatch.setattr(_formatters.openpyxl, "load_workbook", lambda _path: workbook)
    monkeypatch.setattr(workbook, "save", _raise_save_error)

    with pytest.raises(
        FormatterError, match="Failed to save formatted workbook"
    ) as exc_info:
        ExcelFormatter(sample_xlsx).column_width("A", 15).save()

    assert isinstance(exc_info.value.cause, PermissionError)
